# !pip install tflite-runtime
import numpy as np
import tflite_runtime.interpreter as tflite
import openpyxl

excel_location = ""

# Load the TensorFlow Lite model
interpreter = tflite.Interpreter(model_path='//content//kerasModelUpdated1.tflite')
interpreter.allocate_tensors()

# Get input and output details
input_details = interpreter.get_input_details()
output_details = interpreter.get_output_details()

# Load the Excel file and select the active sheet
workbook = openpyxl.load_workbook(excel_location)
sheet = workbook.active

# Get the last row of data from the Excel sheet
last_row = sheet.max_row
last_row_data = []
for col in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=last_row, column=col).value
    last_row_data.append(cell_value)

# Convert the last row data to a NumPy array
input_data = np.array([last_row_data], dtype=np.float32)
print("Input Data from Excel:", input_data)

# Test the model with the input data
interpreter.set_tensor(input_details[0]['index'], input_data)

# Run inference
interpreter.invoke()

# The function `get_tensor()` returns a copy of the tensor data.
output_data = interpreter.get_tensor(output_details[0]['index'])
probability_alz = (output_data[0] * 100).item()  
print("Probability if they have Alzheimer's: {:.2f} %".format(probability_alz))
