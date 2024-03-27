import tkinter
from tkinter import ttk
from tkinter import messagebox
import os
import openpyxl

# Functions Like Get Data Here
def enter_data():
    terms = accept_status_var.get()
    
    if terms == 'Accepted':
        # Identification
        height = height_entry.get()
        weight = weight_entry.get()
        bmi = bmi_entry.get()
        age = age_spinbox.get()
        casi = casi_entry.get()
        mmse = mmse_entry.get()
        

        print("Height: ", height, "Weight: ", weight, "BMI: ", bmi, "Age: ", age)
        print("CASE: ", casi, "MMSE", mmse)
        print("Terms and Conditions: ", terms)
        print("-------------------------------------------------")

        filepath = "//home//seniorDesign//Desktop//project//personData.xlsx"

        if not os.path.exists(filepath):                                                       # Check if the filepath exist if not create it in here
            workbook = openpyxl.Workbook()                                                     # Open workbook like an Excel sheet
            sheet = workbook.worksheets[0]                                                     # workbook.active or The active here is the sheet that is seen at the bottom of excel "sheet 1, sheet 2..."
            sheet.title = "Sheet1"                                                      
            heading = ["Height", "Weight", "BMI", "Age", "CASI", "MMSE", "No of Strides",
                       "Walking Time", "Stride Length", "Stride Frequency", "Stride Speed",
                       "Stride Cadence", "Stride Time", "Stance Time", "Swing Time"]
            sheet.append(heading)                                                              # Write the heading into the excel sheet
            workbook.save(filepath)                                                            # Save it

        # This few lines are for when you after you created the work book and want to save data
        workbook = openpyxl.load_workbook(filepath)
        if not 'Sheet1' in workbook.sheetnames:
            workbook.create_sheet(title = "Sheet1")
            sheet = workbook["Sheet1"]
            heading = ["Height", "Weight", "BMI", "Age", "CASI", "MMSE", "No of Strides",
                       "Walking Time", "Stride Length", "Stride Frequency", "Stride Speed",
                       "Stride Cadence", "Stride Time", "Stance Time", "Swing Time"]
            sheet.append(heading)
            workbook.save(filepath)  

        sheet = workbook["Sheet1"]
        sheet.append([height, weight, bmi, age, casi, mmse])
        workbook.save(filepath)

    else:
        tkinter.messagebox.showwarning(title = "Error", message = "You have not accepted the terms")

# =====================================================================================

# Create a Data Entry Form
window = tkinter.Tk()
window.title("Data Entry Form")
window.geometry("340x280")  # Adjusted width to accommodate three boxes on the left

# Create a canvas to hold the frame
canvas = tkinter.Canvas(window)
canvas.grid(row=0, column=0, sticky="nsew")

# Create a vertical scrollbar
scrollbar = tkinter.Scrollbar(window, orient="vertical", command=canvas.yview)
scrollbar.grid(row=0, column=1, sticky="ns")

# Configure the canvas to use the scrollbar
canvas.configure(yscrollcommand=scrollbar.set)

# Create a frame to hold the widgets
frame = tkinter.Frame(canvas)
canvas.create_window((0, 0), window=frame, anchor="nw")

# Resize the canvas scroll region when the frame size changes
def on_frame_configure(event):
    canvas.configure(scrollregion=canvas.bbox("all"))

frame.bind("<Configure>", on_frame_configure)

# Enable scrolling with the mouse wheel
def on_mousewheel(event):
    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

# Bind the mousewheel event to the canvas
canvas.bind_all("<MouseWheel>", on_mousewheel)

# Add grid_columnconfigure and grid_rowconfigure to the root window
window.grid_columnconfigure(0, weight=1)
window.grid_rowconfigure(0, weight=1)

# =====================================================================================

# User Information
user_info_frame = tkinter.LabelFrame(frame, text = "User Information")
user_info_frame.grid(row = 0, column = 0, sticky = "news", padx = 10, pady = 10)

# First Declare it then put it into the grid
# Height
height_label = tkinter.Label(user_info_frame, text = "Height (CM)")
height_entry = tkinter.Entry(user_info_frame, width = 11)
height_label.grid(row = 0, column = 0)
height_entry.grid(row = 1, column = 0)

# Weight
weight_label = tkinter.Label(user_info_frame, text = "Weight (kilo)")
weight_entry = tkinter.Entry(user_info_frame, width = 11)
weight_label.grid(row = 0, column = 1)
weight_entry.grid(row = 1, column = 1)

# Age
age_label = tkinter.Label(user_info_frame, text = "Age")
age_spinbox = tkinter.Spinbox(user_info_frame, from_ = 18, to = 110, width = 10)
age_label.grid(row = 0, column = 2)
age_spinbox.grid(row = 1, column = 2)

# BMI
bmi_label = tkinter.Label(user_info_frame, text = "BMI")
bmi_entry = tkinter.Entry(user_info_frame, width = 11)
bmi_label.grid(row = 2, column = 0)
bmi_entry.grid(row = 3, column = 0)

# CASI
casi_label = tkinter.Label(user_info_frame, text = "CASI")
casi_entry = tkinter.Entry(user_info_frame, width = 11)
casi_label.grid(row = 2, column = 1)
casi_entry.grid(row = 3, column = 1)

# MMSE
mmse_label = tkinter.Label(user_info_frame, text = "MMSE")
mmse_entry = tkinter.Entry(user_info_frame, width = 11)
mmse_label.grid(row = 2, column = 2)
mmse_entry.grid(row = 3, column = 2)

# Change format here for info frame
for widget in user_info_frame.winfo_children():
    widget.grid_configure(padx = 5, pady = 5)

# ======================================================================================

# Terms & Conditions
terms_con_frame = tkinter.LabelFrame(frame, text = "Terms & Conditions")
terms_con_frame.grid(row = 6, column = 0, sticky = "news", padx = 10, pady = 10)

accept_status_var = tkinter.StringVar(value = "Denied")
accept_check = tkinter.Checkbutton(terms_con_frame, text = "I accept the terms & conditions", variable = accept_status_var, offvalue = "Denied", onvalue = "Accepted")
accept_check.grid(row = 0, column = 0)

# ======================================================================================

# Button to Enter Data
button = tkinter.Button(frame, text = "Enter Data", command = enter_data)
button.grid(row = 7, column = 0, sticky = "news", padx = 10, pady = 10)

window.mainloop()
