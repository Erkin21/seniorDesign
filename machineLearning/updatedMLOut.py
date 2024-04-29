import numpy as np
import tflite_runtime.interpreter as tflite
import openpyxl
import matplotlib.pyplot as plt

excel_location = "//home//seniorDesign//Desktop//project//personData.xlsx"

# Load the TensorFlow Lite model
interpreter = tflite.Interpreter(model_path='//home//seniorDesign//Desktop//machineLearning//kerasModel//kerasModelUpdated1.tflite')
interpreter.allocate_tensors()

# Get input and output details
input_details = interpreter.get_input_details()
output_details = interpreter.get_output_details()

# Load the Excel file and select the active sheet
workbook = openpyxl.load_workbook(excel_location)
sheet = workbook["Sheet1"]

# Get the last row of data from the Excel sheet
last_row = sheet.max_row
last_row_data = []
for col in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=last_row, column=col).value
    last_row_data.append(cell_value)

# Convert the last row data to a NumPy array
input_data = np.array([last_row_data], dtype=np.float32)
print("Input Data from Excel:", input_data)

# Test the model with the input data
interpreter.set_tensor(input_details[0]['index'], input_data)

# Run inference
interpreter.invoke()

# The function `get_tensor()` returns a copy of the tensor data.
output_data = interpreter.get_tensor(output_details[0]['index'])
probability_alz = (output_data[0] * 100).item()  
print("Probability if they have Alzheimer's: {:.2f} %".format(probability_alz))

# Create a new sheet for storing probabilities and graphing
if not 'Alzheimer\'s Probability' in workbook.sheetnames:
    new_sheet = workbook.create_sheet(title='Alzheimer\'s Probability')
    new_sheet.append(['Test #', 'Probability (%)'])
    new_sheet.append([0, 0])
    new_sheet.append([1, probability_alz])  # Start with test number 1
else:
    sheet = workbook['Alzheimer\'s Probability']
    last_test_number = sheet.max_row - 1  # Subtract header row
    next_test_number = last_test_number + 1
    sheet.append([next_test_number, probability_alz])

sheet = workbook['Alzheimer\'s Probability']
# Save the modified workbook
workbook.save(excel_location)

# Plot the probability if there are more than 1 data point
if sheet.max_row > 2:  # Check if there are actual data points (excluding headers)
    test_numbers = []
    probabilities = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        test_numbers.append(row[0].value)
        probabilities.append(row[1].value)

    if len(test_numbers) == 1:
        # If only one data point, use a bar graph
        plt.bar(test_numbers, probabilities)
    else:
        # If multiple data points, use a line graph
        plt.plot(test_numbers, probabilities)
    plt.xlabel('Test #')
    plt.ylabel('Probability of Alzheimer\'s (%)')
    plt.title('Alzheimer\'s Probability over Tests')
    plt.grid(True)
    plt.show()
else:
    print("No data points to plot.")
