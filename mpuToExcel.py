import serial
import openpyxl

serial_port = '/dev/ttyACM0' # could be /dev/ttyACM1
baud_rate = 9600

ser = serial.Serial(serial_port, baud_rate)

# Read a single line from the serial port
data = ser.readline().decode().strip()
print(data)

# Split the data string into individual values
values = data.split(',')

# Open the Excel file
excel_file_path = '//home//seniorDesign//Desktop//project//personData.xlsx'
wb = openpyxl.load_workbook(excel_file_path)
ws = wb.active

# Find the column containing 'No of Strides'
no_of_strides_column = None
for cell in ws[1]:
    if cell.value == 'No of Strides':
        no_of_strides_column = cell.column
        break

if no_of_strides_column is not None:
    # Find the first empty row below the header
    next_row = 2
    while ws.cell(row=next_row, column=no_of_strides_column).value is not None:
        next_row += 1

    # Insert the new data in the first empty row below the header
    for i, value in enumerate(values):
        column = no_of_strides_column + i  # Calculate the column index for each value
        ws.cell(row=next_row, column=column, value=value)

    # Save the changes to the Excel file
    wb.save(excel_file_path)
else:
    print("Header 'No of Strides' not found in the Excel file.")

ser.close()
