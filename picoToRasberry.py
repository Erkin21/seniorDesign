import serial
from openpyxl import load_workbook

# Configure serial communication with Raspberry Pi Pico
ser = serial.Serial('/dev/ttyACM0', 115200, timeout=1)  # Adjust baud rate as needed

# Open the .txt file on Raspberry Pi Pico for reading
ser.write(b'dataMPU6050.txt r\n')  # Adjust filename as needed

# Initialize variable to store the last line
last_line = ''

# Read and transfer the file contents line by line
while True:
    line = ser.readline().decode('utf-8').strip()
    if not line:
        break
    last_line = line  # Store the current line

# Close the serial connection
ser.close()

# Split the last line data
last_line_data = last_line.split()

# Open the Excel file and load the existing sheet
wb = load_workbook('data.xlsx')
sheet = wb['Sheet1']

# Clear existing data in the sheet
for row in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.value = None

# Write each value from the last line to individual cells in Excel
for i, value in enumerate(last_line_data, start=1):
    sheet.cell(row=1, column=i, value=float(value))  # Assuming the data is numeric

# Save the changes to the Excel file
wb.save('data.xlsx')

print("Data written to Excel file successfully.")
